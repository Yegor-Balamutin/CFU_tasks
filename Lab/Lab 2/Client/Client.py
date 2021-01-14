from flask import Flask, request
import openpyxl
import json
import datetime

def data_save_xl(buf):
    print("saving data")
    current_row = sheet.max_row + 1
    for i in range(len(buf)):
        num = find_num(sheet)
        sheet.cell(row=current_row, column=1).value = num
        num += 1
        sheet.cell(row=current_row, column=2).value = buf[i]["user_id"]
        sheet.cell(row=current_row, column=3).value = buf[i]["datetime"]
        for j in range(len(buf[i]["check"])):
            sheet.cell(row=current_row, column=4).value = str(j+1) + ": " + tuple(buf[i]["check"][j].items())[0][0]
            sheet.cell(row=current_row, column=5).value = tuple(buf[i]["check"][j].items())[0][1]
            current_row += 1
    book.save("data.xlsx")

def find_num(sheet):
    num = 1
    for i in range(sheet.max_row, 1, -1):
            if (sheet.cell(row=i, column=1).value != None):
                num = sheet.cell(row=i, column=1).value + 1
                break
    return num

try:
    book = openpyxl.open("data.xlsx", read_only=False)
    print("data.xlsx found")
    sheet = book.active
except FileNotFoundError:
    print("data.xlsx not found")
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.cell(row=1, column=1).value = "N"
    sheet.cell(row=1, column=2).value = "User ID"
    sheet.cell(row=1, column=3).value = "Datetime"
    sheet.cell(row=1, column=4).value = "Item"
    sheet.cell(row=1, column=5).value = "Prise"
    book.save("data.xlsx")

buf = []
 
app = Flask(__name__)
 
@app.route('/', methods=['POST', 'GET'])
def index():
    buf.append(request.json)
    buf[len(buf) - 1]["datetime"] = datetime.datetime.now().ctime()
    if len(buf) > 2:
        data_save_xl(buf)
        buf.clear()
    return "OK"
 
if __name__ == "__main__":
    app.run()

